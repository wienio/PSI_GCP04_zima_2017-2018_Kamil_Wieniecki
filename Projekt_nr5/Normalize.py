from openpyxl import load_workbook
from math import pow
from math import sqrt

### rekordy w excelu ###
### 1-50 ---> setosa
### 51-100 ---> versicolor
### 101-150 ---> virginica

workbook = load_workbook('./Iris.xlsx', data_only=True)
sheet = workbook.get_sheet_names()[0]
worksheet = workbook.get_sheet_by_name(sheet)
text_file = open("flowers.txt", "a")

def normalize(array, number):
    sum = 0
    for i in range(0, len(array)):
        sum += pow(array[i], 2)
    return number / sqrt(sum)

list = []
for row in worksheet.iter_rows():
    for i in range(0, len(row)):
        list.append(row[i].value)
    text_file.write("[ ")
    for i in range(0, 4):
        result = normalize(list, list[i])
        text_file.write(str(result))
        if i != 3:
            text_file.write(", ")
    text_file.write(" ],\n")
    list.clear()

text_file.close()
